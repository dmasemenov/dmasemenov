# -*- coding: utf-8 -*-
"""
Created on Sat Oct  9 12:41:32 2021

@author: HP user
"""


import openpyxl, random
direct = dict()
wb = openpyxl.load_workbook("dict.xlsx")
sh = wb.active
for i in range (1,201):
    voc = sh.cell(row=i, column=1).value
    trans = sh.cell(row=i, column=2).value
    if voc != '':
        direct [voc] = trans
inp = ''
inp2 = ''
while (inp != 'd') and (inp != 'r'):
    print ('Direct or reverse mode [d/r]?')
    inp = input()
if inp == 'd':
    while inp2 != 'n' and direct:
        word, meaning = random.choice(list(direct.items()))
        temp = direct.pop(word)
        print (meaning)#(meaning.strip())
        if input() == word.strip():
            print ('You win!')
        else: 
            print ('Correct word is '+word.strip())
            direct[word]=temp
        print ('Continue [y/n]?')
        inp2 = input()
elif inp == "r":
    while inp2 != 'n' and direct:
        word, meaning = random.choice(list(direct.items()))
        temp = direct.pop(word)
        print (word)
        rand = random.randint(1,4)
        for i in range (1,5):
            outstr = str(i)+'. '
            if i==rand:
                outstr+=meaning.strip()
            else:
                randword, randmeaning = random.choice(list(direct.items()))
                outstr+=randmeaning.strip()
            print (outstr)
        print ('Input number ?')
        if int(input()) == rand:
            print ('Correct answer!')
        else:
            print('Correct answer is '+str(rand))
            direct[word]=temp
        print ('Continue [y/n]?')
        inp2 = input()        
                
            


'''
import json

fin = open('input.json', 'r', encoding='utf8')

dct = json.loads(fin.read())

fin.close()

popup = dct['menu']['popup']

menuitem = popup['menuitem']

for now in menuitem:

    for key in sorted(now):

            print(key, now[key])
'''